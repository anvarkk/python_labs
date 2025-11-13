import csv
from pathlib import Path

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX используя openpyxl.
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Для работы требуется openpyxl. Установите: pip install openpyxl")
    
    # Проверка существования файла
    if not Path(csv_path).exists():
        raise FileNotFoundError(f"CSV файл не найден: {csv_path}")
    
    # Создание директории для выходного файла если не существует
    Path(xlsx_path).parent.mkdir(parents=True, exist_ok=True)
    
    # Создание новой книги Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    
    # Чтение CSV и запись в XLSX
    try:
        with open(csv_path, 'r', encoding='utf-8') as csv_file:
            reader = csv.reader(csv_file)
            rows = list(reader)
    except csv.Error:
        raise ValueError("Неверный формат CSV файла")
    
    if not rows:
        raise ValueError("CSV файл пустой")
    
    # Запись данных
    for row_idx, row in enumerate(rows, 1):
        for col_idx, value in enumerate(row, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Настройка авто-ширины колонок
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Получаем букву колонки
        
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Минимальная ширина - 8 символов
        adjusted_width = max(max_length + 2, 8)
        sheet.column_dimensions[column].width = adjusted_width
    
    # Сохранение файла
    try:
        workbook.save(xlsx_path)
    except Exception as e:
        raise ValueError(f"Ошибка сохранения XLSX: {e}")